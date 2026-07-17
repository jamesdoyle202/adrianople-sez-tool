"""SEZ Polygon Matcher — match master-list points to uploaded GeoJSON polygons."""

from __future__ import annotations

import json
import math
import re
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

import geopandas as gpd
import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from shapely.geometry import Point, shape
from shapely.geometry.base import BaseGeometry

from components.navigation import render_back_link
from config import APP_NAME

LAT_ALIASES = ("lat", "latitude", "y")
LON_ALIASES = ("lon", "lng", "long", "longitude", "x")
POLYGON_NAME_KEYS = ("zone name", "zone_name", "name", "title", "label")
MATCH_COLS = ["Match Type", "Matched File", "Matched Polygon", "Distance (km)"]


def _normalize_header(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value).strip())


def _detect_coordinate_column(columns: List[str], aliases: Tuple[str, ...]) -> Optional[str]:
    normalized = {_normalize_header(column).lower(): column for column in columns}

    for alias in aliases:
        if alias in normalized:
            return normalized[alias]

    for alias in aliases:
        for key, original in normalized.items():
            if key.startswith(alias) or re.search(rf"\b{re.escape(alias)}\b", key):
                return original

    return None


def _read_master_table(uploaded_file) -> pd.DataFrame:
    name = uploaded_file.name.lower()
    if name.endswith((".xlsx", ".xlsm", ".xls")):
        frame = pd.read_excel(uploaded_file)
    elif name.endswith(".tsv"):
        frame = pd.read_csv(uploaded_file, sep="\t")
    else:
        frame = pd.read_csv(uploaded_file)

    frame.columns = [_normalize_header(column) for column in frame.columns]
    return frame


def load_master_list(uploaded_file) -> Tuple[pd.DataFrame, gpd.GeoDataFrame, str, str]:
    frame = _read_master_table(uploaded_file)
    if frame.empty:
        raise ValueError("The master SEZ list is empty.")

    lat_col = _detect_coordinate_column(list(frame.columns), LAT_ALIASES)
    lon_col = _detect_coordinate_column(list(frame.columns), LON_ALIASES)
    if lat_col is None or lon_col is None:
        found = ", ".join(frame.columns.astype(str)) or "(none)"
        raise ValueError(
            "Could not find latitude/longitude columns. "
            f"Looking for lat/latitude/y and lon/lng/long/longitude/x. Columns found: {found}"
        )

    working = frame.copy()
    working["_row_id"] = range(len(working))
    working["_lat"] = pd.to_numeric(working[lat_col], errors="coerce")
    working["_lon"] = pd.to_numeric(working[lon_col], errors="coerce")
    working["_has_coords"] = working["_lat"].notna() & working["_lon"].notna()

    valid = working[working["_has_coords"]].copy()
    geometry = [
        Point(float(lon), float(lat))
        for lon, lat in zip(valid["_lon"], valid["_lat"])
    ]
    points = gpd.GeoDataFrame(valid, geometry=geometry, crs="EPSG:4326")
    return working, points, lat_col, lon_col


def _polygon_name_from_properties(properties: Dict[str, Any]) -> str:
    if not properties:
        return ""

    lowered = {str(key).strip().lower(): value for key, value in properties.items()}
    for key in POLYGON_NAME_KEYS:
        value = lowered.get(key)
        if value not in (None, ""):
            return str(value)
    return ""


def _geometry_to_features(
    geometry_obj: BaseGeometry,
    source_file: str,
    properties: Optional[Dict[str, Any]] = None,
) -> List[Dict[str, Any]]:
    if geometry_obj is None or geometry_obj.is_empty:
        return []

    geom_type = geometry_obj.geom_type
    if geom_type not in ("Polygon", "MultiPolygon"):
        return []

    props = dict(properties or {})
    return [
        {
            "source_file": source_file,
            "polygon_name": _polygon_name_from_properties(props),
            "geometry": geometry_obj,
        }
    ]


def parse_geojson_data(data: Any, source_file: str) -> List[Dict[str, Any]]:
    records: List[Dict[str, Any]] = []

    if isinstance(data, dict) and data.get("type") == "FeatureCollection":
        for feature in data.get("features") or []:
            if not isinstance(feature, dict):
                continue
            geometry = feature.get("geometry")
            if not geometry:
                continue
            records.extend(
                _geometry_to_features(
                    shape(geometry),
                    source_file,
                    feature.get("properties") or {},
                )
            )
        return records

    if isinstance(data, dict) and data.get("type") == "Feature":
        geometry = data.get("geometry")
        if not geometry:
            return []
        return _geometry_to_features(
            shape(geometry),
            source_file,
            data.get("properties") or {},
        )

    if isinstance(data, dict) and data.get("type") in ("Polygon", "MultiPolygon"):
        return _geometry_to_features(shape(data), source_file, {})

    raise ValueError("Unsupported GeoJSON structure.")


def parse_geojson_bytes(content: bytes, source_file: str) -> List[Dict[str, Any]]:
    return parse_geojson_data(json.loads(content), source_file)


def parse_polygon_workbook(content: bytes, source_file: str) -> List[Dict[str, Any]]:
    frame = pd.read_excel(BytesIO(content), engine="openpyxl")
    frame.columns = [_normalize_header(column) for column in frame.columns]
    geometry_column = next(
        (
            column
            for column in frame.columns
            if column.lower() in ("geojson", "geometry")
        ),
        None,
    )
    if geometry_column is None:
        raise ValueError(
            "Excel polygon files need a 'GeoJSON' or 'Geometry' column."
        )

    records: List[Dict[str, Any]] = []
    for _, row in frame.iterrows():
        value = row.get(geometry_column)
        if pd.isna(value) or not str(value).strip():
            continue

        data = json.loads(str(value))
        row_properties = {
            column: value
            for column, value in row.items()
            if column != geometry_column and not pd.isna(value)
        }

        if isinstance(data, dict) and data.get("type") == "Feature":
            data["properties"] = {
                **row_properties,
                **(data.get("properties") or {}),
            }
        elif isinstance(data, dict) and data.get("type") in (
            "Polygon",
            "MultiPolygon",
        ):
            data = {
                "type": "Feature",
                "properties": row_properties,
                "geometry": data,
            }

        records.extend(parse_geojson_data(data, source_file))

    return records


def load_polygons(uploaded_files) -> gpd.GeoDataFrame:
    records: List[Dict[str, Any]] = []

    for uploaded in uploaded_files:
        try:
            if uploaded.name.lower().endswith(".xlsx"):
                file_records = parse_polygon_workbook(
                    uploaded.getvalue(),
                    uploaded.name,
                )
            else:
                file_records = parse_geojson_bytes(
                    uploaded.getvalue(),
                    uploaded.name,
                )
        except Exception as exc:
            st.warning(f"Skipped {uploaded.name}: {exc}")
            continue

        if not file_records:
            st.warning(f"Skipped {uploaded.name}: no polygon geometries found.")
            continue

        records.extend(file_records)

    if not records:
        raise ValueError("No valid polygons were found in the uploaded files.")

    polygons = gpd.GeoDataFrame(records, crs="EPSG:4326")
    polygons["polygon_id"] = range(len(polygons))
    return polygons


def _corrected_km(mercator_meters: float, latitude: float) -> float:
    lat = max(min(float(latitude), 85.0), -85.0)
    return float(mercator_meters) * math.cos(math.radians(lat)) / 1000.0


def _inflated_search_meters(max_km: float, latitudes: pd.Series) -> float:
    if latitudes.empty:
        return max_km * 1000.0
    max_abs_lat = float(latitudes.abs().max())
    max_abs_lat = min(max_abs_lat, 85.0)
    cos_lat = max(math.cos(math.radians(max_abs_lat)), 1e-6)
    return (max_km * 1000.0) / cos_lat


def match_zones(
    master: pd.DataFrame,
    points: gpd.GeoDataFrame,
    polygons: gpd.GeoDataFrame,
    max_km: float,
) -> pd.DataFrame:
    result = master.copy()
    result["Match Type"] = "NO COORDS"
    result["Matched File"] = ""
    result["Matched Polygon"] = ""
    result["Distance (km)"] = pd.NA

    valid_mask = result["_has_coords"].fillna(False)
    result.loc[valid_mask, "Match Type"] = "NO MATCH"

    if points.empty or polygons.empty:
        return _finalize_result(result, master)

    # Pass 1: points inside polygons
    inside = gpd.sjoin(
        points[["_row_id", "geometry"]],
        polygons[["polygon_id", "source_file", "polygon_name", "geometry"]],
        how="inner",
        predicate="within",
    )
    if not inside.empty:
        inside = inside.sort_values(["_row_id", "polygon_id"]).drop_duplicates(
            subset=["_row_id"],
            keep="first",
        )
        for _, row in inside.iterrows():
            idx = result.index[result["_row_id"] == row["_row_id"]][0]
            result.at[idx, "Match Type"] = "INSIDE"
            result.at[idx, "Matched File"] = row["source_file"]
            result.at[idx, "Matched Polygon"] = row["polygon_name"] or ""
            result.at[idx, "Distance (km)"] = 0.0

    unmatched_ids = set(
        result.loc[result["Match Type"] == "NO MATCH", "_row_id"].tolist()
    )
    unmatched_points = points[points["_row_id"].isin(unmatched_ids)].copy()

    # Pass 2: nearest polygon within threshold (Web Mercator + latitude correction)
    if not unmatched_points.empty and max_km > 0:
        search_meters = _inflated_search_meters(max_km, unmatched_points["_lat"])
        points_m = unmatched_points.to_crs(epsg=3857)
        polygons_m = polygons.to_crs(epsg=3857)

        nearby = gpd.sjoin_nearest(
            points_m[["_row_id", "_lat", "geometry"]],
            polygons_m[["polygon_id", "source_file", "polygon_name", "geometry"]],
            how="left",
            max_distance=search_meters,
            distance_col="_mercator_m",
        )

        if not nearby.empty:
            nearby = nearby[nearby["polygon_id"].notna()].copy()
            if not nearby.empty:
                nearby["_true_km"] = [
                    _corrected_km(meters, lat)
                    for meters, lat in zip(nearby["_mercator_m"], nearby["_lat"])
                ]
                nearby = nearby[nearby["_true_km"] <= max_km].copy()
                if not nearby.empty:
                    nearby = nearby.sort_values(["_row_id", "_true_km", "polygon_id"])
                    nearby = nearby.drop_duplicates(subset=["_row_id"], keep="first")
                    for _, row in nearby.iterrows():
                        idx = result.index[result["_row_id"] == row["_row_id"]][0]
                        result.at[idx, "Match Type"] = "NEARBY"
                        result.at[idx, "Matched File"] = row["source_file"]
                        result.at[idx, "Matched Polygon"] = row["polygon_name"] or ""
                        result.at[idx, "Distance (km)"] = round(float(row["_true_km"]), 2)

    return _finalize_result(result, master)


def _finalize_result(result: pd.DataFrame, master: pd.DataFrame) -> pd.DataFrame:
    original_cols = [column for column in master.columns if not str(column).startswith("_")]
    output = result[original_cols + MATCH_COLS].copy()
    output["Distance (km)"] = output["Distance (km)"].astype("Float64")
    return output


def build_excel_bytes(results: pd.DataFrame, max_km: float) -> bytes:
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        results.to_excel(writer, sheet_name="Matches", index=False)

        summary = (
            results["Match Type"]
            .value_counts()
            .reindex(["INSIDE", "NEARBY", "NO MATCH", "NO COORDS"], fill_value=0)
            .reset_index()
        )
        summary.columns = ["Match Type", "Count"]
        summary.to_excel(writer, sheet_name="Summary", index=False, startrow=2)

        matches_sheet = writer.sheets["Matches"]
        summary_sheet = writer.sheets["Summary"]

        header_fill = PatternFill("solid", fgColor="111111")
        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        body_font = Font(name="Arial", size=10)

        for cell in matches_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")

        for row in matches_sheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font

        matches_sheet.freeze_panes = "A2"
        for column_cells in matches_sheet.columns:
            values = [str(cell.value) if cell.value is not None else "" for cell in column_cells]
            width = min(max((len(value) for value in values), default=10) + 2, 45)
            matches_sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width

        summary_sheet["A1"] = f"SEZ Polygon Matcher summary — nearby threshold {max_km:g} km"
        summary_sheet["A1"].font = Font(name="Arial", size=10, bold=True)
        for cell in summary_sheet[3]:
            cell.fill = header_fill
            cell.font = header_font
        for row in summary_sheet.iter_rows(min_row=4, max_row=3 + len(summary)):
            for cell in row:
                cell.font = body_font
        summary_sheet.column_dimensions["A"].width = 18
        summary_sheet.column_dimensions["B"].width = 12

    return buffer.getvalue()


def render() -> None:
    render_back_link(target="home")
    st.markdown(f"# {APP_NAME}")
    st.markdown("### SEZ Polygon Matcher")
    st.caption(
        "Upload a master SEZ list and polygon GeoJSON files. "
        "Match each zone point to a polygon, then download a formatted Excel workbook."
    )
    st.markdown("<br>", unsafe_allow_html=True)

    st.markdown("#### 1 · Master SEZ list")
    master_file = st.file_uploader(
        "Master SEZ list",
        type=["xlsx", "xlsm", "xls", "csv", "tsv"],
        key="matcher_master_file",
        label_visibility="collapsed",
    )

    st.markdown("#### 2 · Polygon files")
    polygon_files = st.file_uploader(
        "Polygon files",
        type=["geojson", "json", "xlsx"],
        accept_multiple_files=True,
        key="matcher_polygon_files",
        label_visibility="collapsed",
    )

    st.markdown("#### 3 · Settings")
    max_km = st.number_input(
        "Nearby threshold (km)",
        min_value=0.0,
        max_value=100.0,
        value=10.0,
        step=1.0,
    )

    st.markdown("<br>", unsafe_allow_html=True)
    ready = master_file is not None and bool(polygon_files)
    run = st.button("Match zones", disabled=not ready, use_container_width=True)

    if not run:
        if not ready:
            st.caption("Upload a master list and at least one polygon file to continue.")
        return

    try:
        with st.spinner("Loading master SEZ list…"):
            master, points, lat_col, lon_col = load_master_list(master_file)

        with st.spinner("Loading polygon files…"):
            polygons = load_polygons(polygon_files)

        with st.spinner("Matching zones…"):
            results = match_zones(master, points, polygons, float(max_km))
    except ValueError as exc:
        st.error(str(exc))
        return
    except Exception as exc:
        st.error(f"Something went wrong: {exc}")
        return

    st.caption(f"Detected coordinates: {lat_col}, {lon_col}")
    st.markdown("<br>", unsafe_allow_html=True)

    counts = results["Match Type"].value_counts()
    col1, col2, col3, col4 = st.columns(4)
    col1.metric("Inside", int(counts.get("INSIDE", 0)))
    col2.metric("Nearby", int(counts.get("NEARBY", 0)))
    col3.metric("No match", int(counts.get("NO MATCH", 0)))
    col4.metric("No coords", int(counts.get("NO COORDS", 0)))

    st.markdown("<br>", unsafe_allow_html=True)
    st.dataframe(results, use_container_width=True)

    excel_bytes = build_excel_bytes(results, float(max_km))
    st.download_button(
        label="Download matches (.xlsx)",
        data=excel_bytes,
        file_name="sez_polygon_matches.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key="matcher_download",
    )
